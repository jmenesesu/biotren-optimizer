"""Parser de demanda — matrices OD por franja (18-03-2026) y perfil de carga.

Convierte:
  - Las 3 matrices OD por franja horaria a formato tidy
    (origen, destino, franja, viajes).
  - El perfil de carga por servicio (afluencia por servicio y sentido).

Uso:
    python parsers/parse_od.py
Salida:
    datos/clean/od_franjas.csv
    datos/clean/perfil_carga.csv
"""
import re
import pandas as pd
import openpyxl
from config import PERFIL_DIR, CLEAN, CAP_AUTOMOTOR

FRANJAS = {
    "[0500 - 1000[": "05-10",
    "[1000 - 1600[": "10-16",
    "[1600 - 2359[": "16-24",
}


def parse_matrices_od():
    registros = []
    for sufijo, franja in FRANJAS.items():
        f = PERFIL_DIR / f"Matriz OD 18032026 {sufijo}.xlsx"
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Export"]
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
        # Fila 3 (índice 2) = encabezados de destino desde la col B
        destinos = [c for c in filas[2][1:] if c]
        for fila in filas[3:]:
            origen = fila[0]
            if not origen or str(origen).lower().startswith("total"):
                continue
            for j, dest in enumerate(destinos, start=1):
                val = fila[j] if j < len(fila) else None
                if val is None or val == "":
                    continue
                registros.append({
                    "origen": str(origen).strip(),
                    "destino": str(dest).strip(),
                    "franja": franja,
                    "viajes": float(val),
                })
    df = pd.DataFrame(registros)
    df.to_csv(CLEAN / "od_franjas.csv", index=False)
    return df


def parse_perfil():
    f = PERFIL_DIR / "Perfil de Carga 18 marzo 2026.xlsx"
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb["Hoja2"]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    # Bloques de 2 columnas por sentido: (servicio, afluencia) en cols 0-1, 3-4, 6-7, 9-10
    sentidos = {0: "CC->CW", 3: "CW->CC", 6: "HQ->TH", 9: "TH->HQ"}
    registros = []
    for fila in filas[2:]:  # saltar las 2 filas de encabezado
        for col, sentido in sentidos.items():
            serv = fila[col] if col < len(fila) else None
            afl = fila[col + 1] if col + 1 < len(fila) else None
            if serv is None or afl is None:
                continue
            s = str(serv).strip()
            if not re.fullmatch(r"\d+", s):  # ignora 'Total general', vacíos
                continue
            registros.append({
                "servicio": s, "sentido": sentido, "afluencia": float(afl),
            })
    df = pd.DataFrame(registros)
    df["supera_capacidad"] = df["afluencia"] > CAP_AUTOMOTOR
    df.to_csv(CLEAN / "perfil_carga.csv", index=False)
    return df


if __name__ == "__main__":
    od = parse_matrices_od()
    perfil = parse_perfil()
    print(f"Pares OD (origen-destino-franja): {len(od)}")
    print("Viajes totales por franja:")
    print(od.groupby("franja")["viajes"].sum().to_string())
    print(f"\nServicios en perfil de carga: {len(perfil)}")
    print(f"Servicios que superan capacidad ({CAP_AUTOMOTOR} pax): "
          f"{int(perfil['supera_capacidad'].sum())}")
    top = perfil.sort_values("afluencia", ascending=False).head(5)
    print("Top 5 servicios más cargados:")
    print(top.to_string(index=False))
    print(f"\nGuardado en: {CLEAN}")
